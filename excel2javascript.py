import argparse
import re
from collections import defaultdict, deque, namedtuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import js2py
from rich.tree import Tree
from rich import print


def convert_to_js(formula):
    """
    Converts Excel formula syntax to JavaScript syntax.
    
    Args:
        formula (str): A string representing an Excel formula.
    
    Returns:
        str: A string representing the equivalent JavaScript expression.
        
    Notes:
        This function handles various Excel-specific syntax quirks and translates them into JavaScript-compatible syntax.
    """
    # Convert percentages with comma as decimal point to float.
    formula = re.sub(r"(\d+),(\d+)%", lambda m: str(int(m.group(1)) + int(m.group(2)) / 100.0) + "/100", formula)
    
    # Convert regular percentages to float.
    formula = re.sub(r"(\d+)%", lambda m: str(int(m.group(1))) + "/100", formula)
    
    # Correctly convert SUM to sequence of additions.
    formula = re.sub(
        r"SUM\((\w+):(\w+)\)", 
        lambda m: '+'.join(
            f"{get_column_letter(column_index_from_string(m.group(1)[:1]) + i)}{m.group(1)[1:]}" 
            for i in range(column_index_from_string(m.group(2)[:1]) - column_index_from_string(m.group(1)[:1]) + 1)
        ), 
        formula
    )
    
    # Replace MIN and MAX with Math.min and Math.max.
    formula = formula.replace("MIN", "Math.min").replace("MAX", "Math.max")
    
    # Replace every instance of a comma between digits with a dot.
    formula = re.sub(r"(\d+),(\d+)", r"\1.\2", formula)
    
    # Remove $ symbols from cell references.
    formula = formula.replace("$", "")
    
    return formula


def extract_and_convert_all_cells(sheet):
    """
    Extracts and converts all cells in the given Excel sheet to JavaScript.
    
    Args:
        sheet (Worksheet): An OpenPyXL Worksheet object representing the Excel sheet.
    
    Returns:
        tuple: A tuple containing:
            - A dictionary mapping cell references to their JavaScript representations.
            - A dependency graph representing the dependencies between cells.
            - A dictionary mapping cell references to their original formulas.
    """
    all_cells = defaultdict(str)
    dependency_graph = defaultdict(set)
    original_formulas = defaultdict(str)
    defined_cells = set()
    
    for row_idx, row in enumerate(sheet.iter_rows(), start=1):
        for col_idx, cell in enumerate(row, start=1):
            cell_ref = get_column_letter(col_idx) + str(row_idx)

            if isinstance(cell.value, (int, float)):
                all_cells[cell_ref] = f"var {cell_ref} = {cell.value};"
                defined_cells.add(cell_ref)
            elif isinstance(cell.value, str) and cell.value.startswith('='):
                original_formulas[cell_ref] = cell.value
                formula = cell.value[1:]
                js_formula = convert_to_js(formula)
                all_cells[cell_ref] = f"var {cell_ref} = {js_formula};"
                defined_cells.add(cell_ref)
                var_refs = re.findall(r'[A-Z]+\$?\d+', formula)
                for var_ref in var_refs:
                    cleaned_var_ref = var_ref.replace('$', '')
                    dependency_graph[cell_ref].add(cleaned_var_ref)
                                          
    # Initialize undefined variables to 0
    all_vars = set(all_cells.keys()) | set(var for deps in dependency_graph.values() for var in deps)
    undefined_vars = all_vars - defined_cells
    for var in undefined_vars:
        all_cells[var] = f"var {var} = 0;"
    
    return all_cells, dependency_graph, original_formulas


def topological_sort_all_cells(graph, all_cells):
    """
    Performs a topological sort on the given graph.
    
    Args:
        graph (dict): A dictionary representing the dependency graph.
        all_cells (dict): A dictionary mapping cell references to their JavaScript representations.
    
    Returns:
        list: A list of cell references in topologically sorted order.
        
    Raises:
        ValueError: If a cycle is detected in the graph.
    """
    sorted_nodes = []
    indegree = defaultdict(int)

    # Compute the in-degree of each node.
    for node, neighbors in graph.items():
        for neighbor in neighbors:
            indegree[neighbor] += 1
    
    # Nodes that have no incoming edges (indegree is zero) can be processed.
    # Initially, all nodes with zero in-degree are added to the queue.
    zero_indegree_nodes = set(node for node in all_cells if indegree[node] == 0)
    queue = deque(zero_indegree_nodes)

    while queue:
        node = queue.popleft()
        sorted_nodes.append(node)
        
        # For each neighbor of the node, reduce its in-degree by 1, 
        # since we are processing the current node and ‘removing’ its outgoing edges.
        for neighbor in graph.get(node, []):
            indegree[neighbor] -= 1
            
            # If in-degree of the neighbor becomes zero, add it to the queue.
            if indegree[neighbor] == 0:
                queue.append(neighbor)

    # If there are still nodes left to process, it means there's a cycle.
    # In this case, we add a node with the smallest in-degree to the queue.
    remaining_nodes = set(all_cells) - set(sorted_nodes)
    while remaining_nodes:
        min_indegree_node = min(remaining_nodes, key=lambda node: indegree[node])
        sorted_nodes.append(min_indegree_node)
        remaining_nodes.remove(min_indegree_node)
        for neighbor in graph.get(min_indegree_node, []):
            indegree[neighbor] -= 1
            
    return sorted_nodes


def resolve_and_sort(graph, all_cells):
    """
    Resolves and sorts all cells based on their dependencies.
    
    Args:
        graph (dict): A dictionary representing the dependency graph.
        all_cells (dict): A dictionary mapping cell references to their JavaScript representations.
        
    Returns:
        list: A list of cell references in resolved and sorted order.
        
    Raises:
        ValueError: If a cycle is detected in the graph.
        
    Notes:
        This function uses a depth-first search approach to resolve dependencies recursively,
        ensuring each cell is resolved before its dependents.
    """
    resolved = set()  # Resolved nodes
    unresolved = set()  # Nodes which are being resolved
    sorted_nodes = []  # The final sorted list of nodes
    
    def resolve(node):
        if node in resolved:
            return True
        if node in unresolved:
            return False  # Found a cycle
        unresolved.add(node)
        for dep in graph.get(node, []):
            if not resolve(dep):
                return False  # Propagate cycle detection
        resolved.add(node)
        unresolved.remove(node)
        sorted_nodes.append(node)
        return True
    
    for node in all_cells:
        if not resolve(node):
            raise ValueError(f"The graph has a cycle, possibly due to a circular reference involving {node}")
    
    return sorted_nodes


def detect_and_break_cycles(graph):
    """
    Detects and breaks any cycles present in the given graph.
    
    Args:
        graph (dict): A dictionary representing the dependency graph.
        
    Notes:
        This function modifies the input graph in-place by removing edges to break detected cycles.
    """
    State = namedtuple('State', 'WHITE GRAY BLACK')
    states = defaultdict(lambda: State.WHITE)
    cycle_edges = set()
    
    def dfs(vertex, path):
        states[vertex] = State.GRAY
        for neighbor in graph[vertex]:
            if states[neighbor] == State.WHITE:
                if dfs(neighbor, path + [vertex]):
                    cycle_edges.add((vertex, neighbor))
            elif states[neighbor] == State.GRAY:
                cycle_edges.add((vertex, neighbor))
                return True  # A cycle is found.
        states[vertex] = State.BLACK
        return False  # No cycle is found.
    
    # Here we are making a copy of the keys (vertices) of the graph 
    # so that we do not modify the graph while iterating over it
    for vertex in list(graph.keys()):
        if states[vertex] == State.WHITE:
            dfs(vertex, [])
    
    # After all vertices have been processed, remove the cycle edges from the graph
    for edge in cycle_edges:
        graph[edge[0]].remove(edge[1])  # Break the cycle by removing the edge.


def convert_excel_to_js(excel_path):
    """
    Converts the given Excel file to JavaScript syntax.
    
    Args:
        excel_path (str): The path to the Excel file.
    
    Returns:
        tuple: A tuple containing:
            - A string representing the JavaScript code equivalent to the Excel file.
            - A dictionary mapping cell references to their original formulas in the Excel file.
    """
    workbook = load_workbook(excel_path, data_only=False)
    sheet = workbook.active
    
    all_cells_js, dependency_graph, original_formulas = extract_and_convert_all_cells(sheet)
    
    # Detect and break any cycles in the dependency graph
    detect_and_break_cycles(dependency_graph)
    
    sorted_all_cells = resolve_and_sort(dependency_graph, all_cells_js)
    sorted_all_js_lines = [all_cells_js[cell_ref] for cell_ref in sorted_all_cells if cell_ref in all_cells_js]
    
    return '\n'.join(sorted_all_js_lines), original_formulas


def execute_js_and_compute_cell(js_code, cell):
    """
    Executes the given JavaScript code and computes the value of the specified cell.
    
    Args:
        js_code (str): The JavaScript code to execute.
        cell (str): The cell whose value to compute.
    
    Returns:
        Any: The computed value of the cell, or None if an error occurs during execution.
    """
    try:
        context = js2py.EvalJs()
        context.execute(js_code)
        return getattr(context, cell, None)
    except js2py.base.PyJsException as e:
        print(f"Error computing {cell}: {str(e)}")
        return None


# def show_dependencies(graph, start_cell, js_code):
#     """
#     Constructs and prints a dependency tree for the given cell using the rich library.
    
#     Args:
#         graph (dict): A dictionary representing the dependency graph.
#         start_cell (str): The cell for which to construct the dependency tree.
#         js_code (str): The JavaScript code generated to compute the values of the cells.
#     """
#     def extract_formula(cell):
#         match = re.search(f"var {cell} = (.*?);", js_code)
#         if match:
#             return match.group(1)
#         return ""
    
#     def format_node(cell, formula):
#         value = execute_js_and_compute_cell(js_code, cell)
#         if formula.replace(".", "", 1).isdigit():  # Check if the formula is a numeric value
#             return f"[magenta]{cell}[/magenta] ({value})"
#         else:
#             return f"[magenta]{cell}[/magenta] ({formula} => {value})"
    
#     formula = extract_formula(start_cell)
#     tree = Tree(format_node(start_cell, formula))
    
#     def build_tree(node, parent):
#         """Recursive function to build the dependency tree."""
#         for dependent in graph.get(node, []):
#             dependent_formula = extract_formula(dependent)
#             formatted_node = format_node(dependent, dependent_formula)
#             branch = parent.add(formatted_node)
#             build_tree(dependent, branch)
    
#     build_tree(start_cell, tree)
#     print(tree)


def build_tree(graph, node, parent, formatter, visited):
    """Recursive function to build the tree."""
    if node in visited:
        print(f"Warning: Detected a circular dependency involving {node}")
        return
    visited.add(node)
    for neighbor in graph.get(node, []):
        branch = parent.add(formatter(neighbor))
        build_tree(graph, neighbor, branch, formatter, visited)
    visited.remove(node)


def show_dependencies(graph, start_cell, js_code):
    """Prints a dependency tree for the given cell using the rich library."""
    def formatter(cell):
        formula = extract_formula(js_code, cell)
        value = execute_js_and_compute_cell(js_code, cell)
        if formula.replace(".", "", 1).isdigit():
            return f"[magenta]{cell}[/magenta] ({value})"
        else:
            return f"[magenta]{cell}[/magenta] ({formula} => {value})"
    
    if start_cell:
        tree = Tree(formatter(start_cell))
        build_tree(graph, start_cell, tree, formatter, set())
        print(tree)
    else:
        roots = set(graph.keys()) - set(dependent for dependents in graph.values() for dependent in dependents)
        for node in roots:
            tree = Tree(formatter(node))
            build_tree(graph, node, tree, formatter, set())
            print(tree)


def show_dependants(graph, start_cell, js_code):
    """Prints a dependant tree for the given cell using the rich library."""
    reversed_graph = reverse_graph(graph)
    
    def formatter(cell):
        formula = extract_formula(js_code, cell)
        value = execute_js_and_compute_cell(js_code, cell)
        if formula.replace(".", "", 1).isdigit():
            return f"[magenta]{cell}[/magenta] ({value})"
        else:
            return f"[magenta]{cell}[/magenta] ({formula} => {value})"
    
    if start_cell:
        tree = Tree(formatter(start_cell))
        for neighbor in reversed_graph.get(start_cell, []):
            branch = tree.add(formatter(neighbor))
        print(tree)
    else:
        # If no start cell is specified, show the dependants for all root nodes.
        roots = set(reversed_graph.keys()) - set(dependant for dependants in reversed_graph.values() for dependant in dependants)
        for node in roots:
            tree = Tree(formatter(node))
            for neighbor in reversed_graph.get(node, []):
                branch = tree.add(formatter(neighbor))
            print(tree)


def extract_formula(js_code, cell):
    match = re.search(f"var {cell} = (.*?);", js_code)
    if match:
        return match.group(1)
    return ""


def reverse_graph(graph):
    """Reverses the direction of the graph edges."""
    reversed_graph = {}
    for node, dependents in graph.items():
        for dependent in dependents:
            if dependent in reversed_graph:
                reversed_graph[dependent].add(node)
            else:
                reversed_graph[dependent] = set([node])
    return reversed_graph


if __name__ == "__main__":
    """
    Parses command-line arguments and executes the appropriate actions based on them.
    This can include converting an Excel file to JavaScript, computing the value of a specific cell, 
    printing the original formula or value of a specific cell, or writing the generated JavaScript to a file.
    """
    parser = argparse.ArgumentParser(description="Convert Excel to JS")
    parser.add_argument('excel_file', help='Path to Excel file')
    parser.add_argument('-c', '--compute', help='Compute the value of a specific cell using generated JS')
    parser.add_argument('-f', '--formula', help='Print the formula or numeric value of a specific cell from Excel')
    parser.add_argument('-o', '--output', help='Path to JS file to output to, optional, if not provided, will print to stdout')
    parser.add_argument('-d', '--show-dependencies', nargs='?', const=None, help='Show the dependency tree of a specific cell or of all cells if no cell is specified')
    parser.add_argument('-s', '--show-dependants', help='Show the dependant tree of a specific cell')
    args = parser.parse_args()

    generated_js, original_formulas = convert_excel_to_js(args.excel_file)
    _, dependency_graph, _ = extract_and_convert_all_cells(load_workbook(args.excel_file, data_only=False).active)

    if args.show_dependencies is not None:
        show_dependencies(dependency_graph, args.show_dependencies, generated_js)
    elif args.show_dependants:
        show_dependants(dependency_graph, args.show_dependants, generated_js)
    elif args.formula:
        print(f"The original formula/value of {args.define} is {original_formulas.get(args.define, 'Not Found')}")
    elif args.compute:
        computed_value = execute_js_and_compute_cell(generated_js, args.compute)
        print(f"The computed value of {args.compute} is {computed_value}")
    elif args.output:
        # Test if the JavaScript contains any errors
        test_cell = list(original_formulas.keys())[0] if original_formulas else None
        if test_cell:
            computed_value = execute_js_and_compute_cell(generated_js, test_cell)
            if computed_value is None:
                print(f"Error in the generated JavaScript. Not saving to {args.output}.")
                exit(1)
        
        # Try writing the JavaScript to the specified output file
        try:
            with open(args.output, 'w') as js_file:
                js_file.write(generated_js)
                print(f"Successfully saved JavaScript to {args.output}")
        except IOError as e:
            print(f"Error writing to {args.output}: {str(e)}")
    else:
        print(generated_js)

